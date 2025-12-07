import os.path
import configparser
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
import openpyxl
from openpyxl import Workbook
from shopee_module import ShopeeAutomation

# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/drive"]

def get_gdrive_service():
    """
    Authenticates with the Google Drive API and returns a service object.
    """
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # Make sure you have the credentials.json file from Google Cloud Console
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES, redirect_uri='urn:ietf:wg:oauth:2.0:oob'
            )
            # Use manual authorization flow
            auth_url, _ = flow.authorization_url(prompt='consent')
            
            print("\n" + "="*70)
            print("AUTHORIZATION REQUIRED")
            print("="*70)
            print("\nPlease visit this URL to authorize this application:")
            print("\n" + auth_url + "\n")
            print("After authorization, you will get a code.")
            code = input("Enter the authorization code here: ").strip()
            flow.fetch_token(code=code)
            creds = flow.credentials
        
        # Save the credentials for the next run
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    try:
        service = build("drive", "v3", credentials=creds)
        print("Successfully connected to Google Drive API.")
        return service
    except HttpError as error:
        print(f"An error occurred: {error}")
        return None

def load_config():
    """Load configuration from config.ini file."""
    config = configparser.ConfigParser()
    config.read('config.ini')
    return config

def upload_to_gdrive(service, file_path, folder_id):
    """
    Uploads a file to Google Drive and returns the shareable link.
    
    Args:
        service: Google Drive API service object
        file_path: Path to the file to upload
        folder_id: ID of the Google Drive folder to upload to
    
    Returns:
        str: Shareable link to the uploaded file, or None if upload fails
    """
    try:
        file_name = os.path.basename(file_path)
        file_metadata = {
            'name': file_name,
            'parents': [folder_id]
        }
        
        media = MediaFileUpload(file_path, resumable=True)
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id, webViewLink'
        ).execute()
        
        # Make the file accessible to anyone with the link
        permission = {
            'type': 'anyone',
            'role': 'reader'
        }
        service.permissions().create(
            fileId=file.get('id'),
            body=permission
        ).execute()
        
        print(f"✓ Uploaded: {file_name}")
        return file.get('webViewLink')
        
    except HttpError as error:
        print(f"✗ Error uploading {file_path}: {error}")
        return None

def create_excel_report(order_data, output_file='shopee_report.xlsx'):
    """
    Creates an Excel report with order numbers and Google Drive links.
    Format follows Shopee CS template with 3 columns: No, OrderSN, Bukti
    If the file exists, it will append new data. Otherwise, create new file.
    
    Args:
        order_data: List of dictionaries with 'order_number' and 'gdrive_link'
        output_file: Name of the output Excel file
    
    Returns:
        str: Path to the created Excel file
    """
    # Check if file already exists
    if os.path.exists(output_file):
        print(f"File '{output_file}' sudah ada, akan menambahkan data baru...")
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Find the next empty row (skip header row)
        next_row = ws.max_row + 1
        
        # Get current number for sequential numbering
        current_no = next_row - 1  # Assuming row 1 is header
        
        # Add new data
        for data in order_data:
            ws[f'A{next_row}'] = current_no  # Sequential number
            ws[f'B{next_row}'] = data['order_number']
            ws[f'C{next_row}'] = data['gdrive_link']
            next_row += 1
            current_no += 1
        
        print(f"✓ Menambahkan {len(order_data)} data baru ke file existing")
    else:
        print(f"Membuat file Excel baru: {output_file}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Add headers (Row 1) - matching Shopee template
        ws['A1'] = "No"
        ws['B1'] = "OrderSN/ Nomor Pesanan"
        ws['C1'] = "Bukti pembeli sudah menerima pesanan\n- Screenshot yang menunjukkan pembeli sudah mengonfirmasi menerima produk non fisik. Screenshot harus dari Chat di Shopee, screenshot dari platform lain (cth Whatsapp) tidak akan diproses\n- Masukkan foto kedalam google drive dan salin ulang link kedalam kolom dibawah ini\n- Pastikan google drive tidak terkunci sehingga dapat diakses oleh Tim Shopee"
        
        # Make headers bold and wrap text
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = openpyxl.styles.Font(bold=True)
            ws[cell].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        # Add data starting from row 2
        for idx, data in enumerate(order_data, start=2):
            ws[f'A{idx}'] = idx - 1  # Sequential number starting from 1
            ws[f'B{idx}'] = data['order_number']
            ws[f'C{idx}'] = data['gdrive_link']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 100
        
        # Set row height for header
        ws.row_dimensions[1].height = 80
    
    wb.save(output_file)
    print(f"✓ Excel report created/updated: {output_file}")
    return output_file

def main():
    """Main function to run the full automation workflow."""
    print("\n" + "="*70)
    print("SHOPEE AUTOMATION - FULL WORKFLOW")
    print("="*70)
    
    # Load configuration
    config = load_config()
    folder_id = config.get('GOOGLE_DRIVE', 'FOLDER_ID')
    username = config.get('SHOPEE', 'USERNAME')
    password = config.get('SHOPEE', 'PASSWORD')
    chrome_profile = config.get('SHOPEE', 'CHROME_PROFILE', fallback='Default')
    
    # Check if credentials are configured
    if username == 'your_shopee_username' or password == 'your_shopee_password':
        print("\n✗ ERROR: Shopee credentials not configured!")
        print("Please edit config.ini and add your Shopee username and password.")
        return
    
    # Step 1: Connect to Google Drive
    print("\n[1/5] Connecting to Google Drive...")
    gdrive_service = get_gdrive_service()
    if not gdrive_service:
        print("✗ Could not connect to Google Drive. Aborting.")
        return
    print("✓ Google Drive connected!")
    
    # Step 2: Initialize Shopee automation
    print("\n[2/5] Initializing Shopee automation...")
    shopee = ShopeeAutomation(username, password, headless=False, chrome_profile=chrome_profile)
    shopee.start_browser()
    
    try:
        # Step 3: Login to Shopee
        print("\n[3/5] Logging in to Shopee Seller Centre...")
        if not shopee.login():
            print("✗ Login failed. Aborting.")
            return
        
        # Step 4: Get orders and process
        print("\n[4/5] Getting orders and taking screenshots...")
        order_numbers = shopee.get_orders_to_ship()
        
        if not order_numbers:
            print("\n⚠ No orders found. Would you like to manually enter order numbers?")
            response = input("Enter 'y' to manually input orders, or any key to exit: ").strip().lower()
            if response == 'y':
                print("\nEnter order numbers (one per line, press Enter twice when done):")
                order_numbers = []
                while True:
                    order = input().strip()
                    if not order:
                        break
                    order_numbers.append(order)
            
            if not order_numbers:
                print("No orders to process. Exiting.")
                return
        
        # Process each order
        order_data = []
        screenshots_folder = 'screenshots'
        
        print(f"\nProcessing {len(order_numbers)} orders...")
        for i, order_number in enumerate(order_numbers, 1):
            print(f"\n[{i}/{len(order_numbers)}] Processing order: {order_number}")
            
            # Take screenshot
            screenshot_path = shopee.take_chat_screenshot(order_number, screenshots_folder)
            
            if screenshot_path:
                # Upload to Google Drive
                print(f"  → Uploading to Google Drive...")
                gdrive_link = upload_to_gdrive(gdrive_service, screenshot_path, folder_id)
                
                if gdrive_link:
                    order_data.append({
                        'order_number': order_number,
                        'gdrive_link': gdrive_link
                    })
                    print(f"  ✓ Order {order_number} processed successfully!")
                else:
                    print(f"  ✗ Failed to upload screenshot for order {order_number}")
            else:
                print(f"  ✗ Failed to take screenshot for order {order_number}")
        
        # Step 5: Generate Excel report
        print("\n[5/5] Generating Excel report...")
        if order_data:
            excel_file = create_excel_report(order_data, 'shopee_report.xlsx')
            print(f"\n✓ Excel report created: {excel_file}")
            
            print("\n" + "="*70)
            print("✓ AUTOMATION COMPLETED SUCCESSFULLY!")
            print("="*70)
            print(f"\nSummary:")
            print(f"  - Orders processed: {len(order_data)}")
            print(f"  - Screenshots uploaded to Google Drive")
            print(f"  - Excel report: shopee_report.xlsx")
            print(f"\nNext steps:")
            print(f"  1. Open shopee_report.xlsx")
            print(f"  2. Verify all data is correct")
            print(f"  3. Submit the report to Shopee CS")
        else:
            print("\n⚠ No orders were successfully processed.")
            
    except KeyboardInterrupt:
        print("\n\n⚠ Process interrupted by user.")
    except Exception as e:
        print(f"\n✗ Error during automation: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # Cleanup
        print("\nClosing browser...")
        shopee.close_browser()
        print("✓ Automation finished.")


if __name__ == "__main__":
    main()

